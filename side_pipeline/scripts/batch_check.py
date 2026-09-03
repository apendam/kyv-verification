#!/usr/bin/env python3
"""Run a list of side/axle uploads through the gate sequence, one by one, and
write the results to an Excel file -- one row per image, columns in this
order: the human-provided claim, then the script's own findings in the same
sequential order the checks run, then final decision, then cost/latency/
token totals.

Manifest columns (CSV or .xlsx, header row required):
    image_path             - required. Path to the image (relative paths
                              resolve against the manifest's own directory).
    claimed_vehicle_type    - required. "bus" or "truck".
    claimed_axle_count      - required. Integer.
    claimed_vrn             - required. Claimed vehicle registration number.
    front_image_upload_id   - optional. Audit link to the corresponding
                              front-image record -- not used by any check.
    view_type               - optional. "pure_side" | "corner" -- captured
                              only, doesn't change any check.
    upload_id               - optional. Auto-generated if left blank.
    vision_model            - optional. Per-row model override; falls back
                              to --model, then config.DEFAULT_VISION_MODEL.

Usage:
    python scripts/batch_check.py --manifest uploads.csv --output results.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv  # noqa: E402

load_dotenv()

from openpyxl import Workbook  # noqa: E402

from kyv_side import config, db  # noqa: E402
from kyv_side.client import OpenRouterClient, OpenRouterInsufficientCredits  # noqa: E402
from kyv_side.side_sequence import run_side_sequence  # noqa: E402

REQUIRED_MANIFEST_COLUMNS = ("image_path", "claimed_vehicle_type", "claimed_axle_count", "claimed_vrn")

OUTPUT_COLUMNS = [
    "image_path", "upload_id", "claimed_vehicle_type", "claimed_axle_count", "claimed_vrn",
    "front_image_upload_id", "view_type", "vision_model",
    "detected_vehicle_type", "vehicle_type_match", "is_altered_or_ai_generated", "type_tamper_check_reasoning",
    "full_side_visible", "detected_axle_count", "axle_count_match",
    "vrn_visible", "vrn_readable", "vrn_conflicting_renderings", "vrn_value_read", "vrn_fuzzy_outcome",
    "duplicate_signal_used", "is_duplicate", "duplicate_match_upload_ids",
    "final_decision", "final_reason",
    "latency_ms_vision_checks", "latency_ms_total",
    "prompt_tokens", "completion_tokens", "cost_usd",
]

_VISION_CHECK_NAMES = {"type_tamper_check", "framing_check", "axle_count_check", "vrn_read"}


def read_manifest(path: str) -> list[dict]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        with open(path, newline="") as f:
            rows = [dict(row) for row in csv.DictReader(f)]
    elif ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() for h in next(rows_iter)]
        rows = [dict(zip(header, row)) for row in rows_iter if any(c is not None for c in row)]
    else:
        raise ValueError(f"unsupported manifest format {ext!r} -- use .csv or .xlsx")

    missing = set(REQUIRED_MANIFEST_COLUMNS) - rows[0].keys() if rows else set()
    if missing:
        raise ValueError(f"manifest is missing required column(s): {sorted(missing)}")
    return rows


def build_row(manifest_row: dict, *, image_path: str, upload_id: str, vision_model: str,
              decision: str, reason: str, checks: list[dict]) -> dict:
    by_name: dict[str, dict] = {}
    for c in checks:
        by_name[c["check_name"]] = c  # later rows win on a --force re-run

    row: dict = {
        "image_path": image_path,
        "upload_id": upload_id,
        "claimed_vehicle_type": manifest_row["claimed_vehicle_type"],
        "claimed_axle_count": manifest_row["claimed_axle_count"],
        "claimed_vrn": manifest_row["claimed_vrn"],
        "front_image_upload_id": manifest_row.get("front_image_upload_id"),
        "view_type": manifest_row.get("view_type"),
        "vision_model": vision_model,
    }

    type_tamper = by_name.get("type_tamper_check")
    if type_tamper:
        d = type_tamper["detail"]
        detected_type = d.get("detected_vehicle_type")
        row["detected_vehicle_type"] = detected_type
        claimed_type = (manifest_row["claimed_vehicle_type"] or "").strip().lower()
        row["vehicle_type_match"] = detected_type == claimed_type if detected_type else None
        row["is_altered_or_ai_generated"] = d.get("is_altered_or_ai_generated")
        row["type_tamper_check_reasoning"] = d.get("reasoning")
    else:
        row.update(dict.fromkeys(
            ["detected_vehicle_type", "vehicle_type_match", "is_altered_or_ai_generated",
             "type_tamper_check_reasoning"]))

    framing = by_name.get("framing_check")
    row["full_side_visible"] = framing["detail"].get("full_side_visible") if framing else None

    axle = by_name.get("axle_count_check")
    if axle:
        d = axle["detail"]
        detected_axle_count = d.get("detected_axle_count")
        row["detected_axle_count"] = detected_axle_count
        try:
            claimed_axle_count = int(manifest_row["claimed_axle_count"])
        except (TypeError, ValueError):
            claimed_axle_count = None
        row["axle_count_match"] = (detected_axle_count == claimed_axle_count
                                    if detected_axle_count is not None and claimed_axle_count is not None
                                    else None)
    else:
        row.update(dict.fromkeys(["detected_axle_count", "axle_count_match"]))

    vrn_read = by_name.get("vrn_read")
    vrn_check = by_name.get("vrn_check")
    if vrn_read:
        d = vrn_read["detail"]
        row["vrn_visible"] = d.get("vrn_visible")
        row["vrn_readable"] = d.get("vrn_readable")
        row["vrn_conflicting_renderings"] = d.get("conflicting_renderings")
        row["vrn_value_read"] = d.get("vrn_text")
    else:
        row.update(dict.fromkeys(["vrn_visible", "vrn_readable", "vrn_conflicting_renderings", "vrn_value_read"]))
    row["vrn_fuzzy_outcome"] = vrn_check["verdict"] if vrn_check else None

    dup_c = by_name.get("duplicate_check")
    if dup_c:
        d = dup_c["detail"]
        matches = d.get("matches", [])
        row["duplicate_signal_used"] = d.get("signal")
        row["is_duplicate"] = dup_c["verdict"] == "duplicate"
        row["duplicate_match_upload_ids"] = ", ".join(m["upload_id"] for m in matches)
    else:
        row.update(dict.fromkeys(["duplicate_signal_used", "is_duplicate", "duplicate_match_upload_ids"]))

    row["final_decision"] = decision
    row["final_reason"] = reason

    vision_latency = sum(c["latency_ms"] for c in checks if c["check_name"] in _VISION_CHECK_NAMES)
    total_latency = sum(c["latency_ms"] for c in checks)
    row["latency_ms_vision_checks"] = vision_latency
    row["latency_ms_total"] = total_latency
    row["prompt_tokens"] = sum(c["prompt_tokens"] for c in checks)
    row["completion_tokens"] = sum(c["completion_tokens"] for c in checks)
    row["cost_usd"] = sum(c["cost_usd"] for c in checks)

    return row


def write_excel(rows: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Results"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(col) for col in OUTPUT_COLUMNS])
    wb.save(output_path)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--manifest", required=True, help="CSV or .xlsx file listing images to check.")
    ap.add_argument("--output", required=True, help="Excel file to write results to.")
    ap.add_argument("--model", default=config.DEFAULT_VISION_MODEL,
                     help=f"Default vision model for rows without their own 'vision_model' "
                          f"column value (default: {config.DEFAULT_VISION_MODEL}).")
    ap.add_argument("--db", default=str(config.DEFAULT_DB_PATH), help="SQLite file path.")
    ap.add_argument("--force", action="store_true",
                     help="Re-run rows whose upload_id was already checked (by default, skipped).")
    args = ap.parse_args()

    manifest_rows = read_manifest(args.manifest)
    manifest_dir = Path(args.manifest).resolve().parent
    conn = db.connect(args.db)
    client = OpenRouterClient()

    output_rows: list[dict] = []
    for i, manifest_row in enumerate(manifest_rows, start=1):
        image_path = manifest_row["image_path"].strip()
        if not Path(image_path).is_absolute():
            candidate = manifest_dir / image_path
            image_path = str(candidate if candidate.is_file() else Path(image_path))

        vrn = (manifest_row["claimed_vrn"] or "").strip()
        upload_id = (manifest_row.get("upload_id") or "").strip() or f"side-{vrn}-{time.time_ns()}"
        vision_model = (manifest_row.get("vision_model") or "").strip() or args.model

        print(f"[{i}/{len(manifest_rows)}] {upload_id} <- {image_path}", file=sys.stderr)

        if not Path(image_path).is_file():
            print(f"  skipped: image not found: {image_path}", file=sys.stderr)
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision="SKIPPED",
                                          reason="image not found", checks=[]))
            continue

        if not args.force and db.already_checked(conn, upload_id):
            print(f"  skipped: '{upload_id}' already checked (pass --force to re-run)", file=sys.stderr)
            row = conn.execute("SELECT decision, reason FROM results WHERE upload_id = ?",
                                (upload_id,)).fetchone()
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision=row[0], reason=row[1],
                                          checks=db.fetch_checks_for_upload(conn, upload_id)))
            continue

        try:
            claimed_axle_count = int(manifest_row["claimed_axle_count"])
        except (TypeError, ValueError):
            print(f"  error: claimed_axle_count is not an integer: {manifest_row['claimed_axle_count']!r}",
                  file=sys.stderr)
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision="ERROR",
                                          reason="claimed_axle_count is not an integer", checks=[]))
            continue

        try:
            result = run_side_sequence(
                conn, client, image_path=image_path, claimed_vehicle_type=manifest_row["claimed_vehicle_type"],
                claimed_axle_count=claimed_axle_count, claimed_vrn=vrn, upload_id=upload_id,
                vision_model=vision_model,
            )
        except OpenRouterInsufficientCredits as exc:
            print(f"Stopped: {exc}", file=sys.stderr)
            print("Top up your OpenRouter account before retrying -- this isn't a per-image "
                  "failure, every remaining row will fail the same way until you do.", file=sys.stderr)
            break
        except Exception as exc:  # noqa: BLE001 - one bad row shouldn't kill the whole batch
            print(f"  error: {exc}", file=sys.stderr)
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision="ERROR", reason=str(exc),
                                          checks=[]))
            continue

        output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                      vision_model=vision_model, decision=result.decision,
                                      reason=result.reason,
                                      checks=db.fetch_checks_for_upload(conn, upload_id)))

    write_excel(output_rows, args.output)

    t = client.totals
    print(f"\n-- wrote {len(output_rows)} row(s) to {args.output} --", file=sys.stderr)
    print(f"-- this run: {t.calls} call(s), {t.prompt_tokens + t.completion_tokens} tokens, "
          f"${t.cost_usd:.5f} --", file=sys.stderr)


if __name__ == "__main__":
    main()
