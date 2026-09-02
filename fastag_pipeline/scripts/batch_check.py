#!/usr/bin/env python3
"""Run a list of FASTag uploads through the gate sequence, one by one, and
write the results to an Excel file -- one row per image, columns in this
order: the human-provided claim, then the script's own findings in the same
sequential order the checks run, then the branch taken, then the final
decision, then cost/latency/token totals.

Manifest columns (CSV or .xlsx, header row required):
    image_path         - required. Path to the image (relative paths are
                          resolved relative to the manifest's own directory).
    claimed_vrn         - required. Claimed vehicle registration number this
                          FASTag is filed against (audit metadata only --
                          not a gate in this flow).
    claimed_barcode     - required. Claimed barcode number printed on the tag.
    claimed_tag_id      - required. Claimed Tag ID (from the QR).
    claimed_bank_code   - required. Claimed issuing bank/code (from the QR).
    upload_id           - optional. Auto-generated if left blank.
    vision_model        - optional. Per-row model override; falls back to
                          --model, then config.DEFAULT_VISION_MODEL.

Usage:
    python scripts/batch_check.py --manifest uploads.csv --output results.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv  # noqa: E402

load_dotenv()

from openpyxl import Workbook  # noqa: E402

from kyv_fastag import config, db  # noqa: E402
from kyv_fastag.client import OpenRouterClient, OpenRouterInsufficientCredits  # noqa: E402
from kyv_fastag.fastag_sequence import run_fastag_sequence  # noqa: E402

REQUIRED_MANIFEST_COLUMNS = ("image_path", "claimed_vrn", "claimed_barcode", "claimed_tag_id", "claimed_bank_code")

OUTPUT_COLUMNS = [
    "image_path", "upload_id", "claimed_vrn", "claimed_barcode", "claimed_tag_id", "claimed_bank_code",
    "vision_model",
    "fastag_fully_framed", "is_altered_or_ai_generated", "front_image_reasoning",
    "barcode_readable", "barcode_value_read", "qr_readable", "tag_id_value_read", "bank_code_value_read",
    "path_taken", "barcode_match", "barcode_checksum_valid", "qr_tag_id_match", "qr_bank_code_match",
    "final_decision", "final_reason",
    "latency_ms_vision_check", "latency_ms_total",
    "prompt_tokens", "completion_tokens", "cost_usd",
]

_VISION_CHECK_NAMES = {"fastag_front_check"}


def read_manifest(path: str) -> list[dict]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        with open(path, newline="") as f:
            rows = [dict(row) for row in csv.DictReader(f)]
    elif ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() for h in next(rows_iter)]
        rows = [dict(zip(header, row)) for row in rows_iter if any(c is not None for c in row)]
    else:
        raise ValueError(f"unsupported manifest format {ext!r} -- use .csv or .xlsx")

    missing = set(REQUIRED_MANIFEST_COLUMNS) - rows[0].keys() if rows else set()
    if missing:
        raise ValueError(f"manifest is missing required column(s): {sorted(missing)}")
    return rows


def build_row(manifest_row: dict, *, image_path: str, upload_id: str, vision_model: str,
              decision: str, reason: str, path_taken: str | None, checks: list[dict]) -> dict:
    by_name: dict[str, dict] = {}
    for c in checks:
        by_name[c["check_name"]] = c  # later rows win on a --force re-run

    row: dict = {
        "image_path": image_path,
        "upload_id": upload_id,
        "claimed_vrn": manifest_row["claimed_vrn"],
        "claimed_barcode": manifest_row["claimed_barcode"],
        "claimed_tag_id": manifest_row["claimed_tag_id"],
        "claimed_bank_code": manifest_row["claimed_bank_code"],
        "vision_model": vision_model,
    }

    front = by_name.get("fastag_front_check")
    if front:
        d = front["detail"]
        row["fastag_fully_framed"] = d.get("fastag_fully_framed")
        row["is_altered_or_ai_generated"] = d.get("is_altered_or_ai_generated")
        row["front_image_reasoning"] = d.get("reasoning")
    else:
        row.update(dict.fromkeys(["fastag_fully_framed", "is_altered_or_ai_generated", "front_image_reasoning"]))

    readability = by_name.get("barcode_qr_readability")
    if readability:
        d = readability["detail"]
        row["barcode_readable"] = d.get("barcode_readable")
        row["qr_readable"] = d.get("qr_readable")
    else:
        row.update(dict.fromkeys(["barcode_readable", "qr_readable"]))

    barcode_check = by_name.get("barcode_check")
    row["barcode_value_read"] = barcode_check["detail"].get("barcode_value_read") if barcode_check else None
    row["barcode_match"] = barcode_check["verdict"] == "match" if barcode_check else None

    stripe_check = by_name.get("barcode_stripe_validity")
    row["barcode_checksum_valid"] = stripe_check["detail"].get("checksum_valid") if stripe_check else None

    qr_parsing = by_name.get("qr_parsing")
    if qr_parsing:
        d = qr_parsing["detail"]
        row["tag_id_value_read"] = d.get("tag_id_value_read")
        row["bank_code_value_read"] = d.get("bank_code_value_read")
        row["qr_tag_id_match"] = (d.get("tag_id_value_read") or "").strip().upper() == \
            (manifest_row["claimed_tag_id"] or "").strip().upper()
        row["qr_bank_code_match"] = (d.get("bank_code_value_read") or "").strip().upper() == \
            (manifest_row["claimed_bank_code"] or "").strip().upper()
    else:
        row.update(dict.fromkeys(["tag_id_value_read", "bank_code_value_read", "qr_tag_id_match", "qr_bank_code_match"]))

    row["path_taken"] = path_taken
    row["final_decision"] = decision
    row["final_reason"] = reason

    vision_latency = sum(c["latency_ms"] for c in checks if c["check_name"] in _VISION_CHECK_NAMES)
    total_latency = sum(c["latency_ms"] for c in checks)
    row["latency_ms_vision_check"] = vision_latency
    row["latency_ms_total"] = total_latency
    row["prompt_tokens"] = sum(c["prompt_tokens"] for c in checks)
    row["completion_tokens"] = sum(c["completion_tokens"] for c in checks)
    row["cost_usd"] = sum(c["cost_usd"] for c in checks)

    return row


def write_excel(rows: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Results"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(col) for col in OUTPUT_COLUMNS])
    wb.save(output_path)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--manifest", required=True, help="CSV or .xlsx file listing images to check.")
    ap.add_argument("--output", required=True, help="Excel file to write results to.")
    ap.add_argument("--model", default=config.DEFAULT_VISION_MODEL,
                     help=f"Default vision model for rows without their own 'vision_model' "
                          f"column value (default: {config.DEFAULT_VISION_MODEL}).")
    ap.add_argument("--db", default=str(config.DEFAULT_DB_PATH), help="SQLite file path.")
    ap.add_argument("--force", action="store_true",
                     help="Re-run rows whose upload_id was already checked (by default, skipped).")
    args = ap.parse_args()

    manifest_rows = read_manifest(args.manifest)
    manifest_dir = Path(args.manifest).resolve().parent
    conn = db.connect(args.db)
    client = OpenRouterClient()

    output_rows: list[dict] = []
    for i, manifest_row in enumerate(manifest_rows, start=1):
        image_path = manifest_row["image_path"].strip()
        if not Path(image_path).is_absolute():
            candidate = manifest_dir / image_path
            image_path = str(candidate if candidate.is_file() else Path(image_path))

        upload_id = (manifest_row.get("upload_id") or "").strip() or \
            f"fastag-{manifest_row['claimed_tag_id']}-{time.time_ns()}"
        vision_model = (manifest_row.get("vision_model") or "").strip() or args.model

        print(f"[{i}/{len(manifest_rows)}] {upload_id} <- {image_path}", file=sys.stderr)

        if not Path(image_path).is_file():
            print(f"  skipped: image not found: {image_path}", file=sys.stderr)
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision="SKIPPED",
                                          reason="image not found", path_taken=None, checks=[]))
            continue

        if not args.force and db.already_checked(conn, upload_id):
            print(f"  skipped: '{upload_id}' already checked (pass --force to re-run)", file=sys.stderr)
            row = conn.execute("SELECT decision, reason, path_taken FROM results WHERE upload_id = ?",
                                (upload_id,)).fetchone()
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision=row[0], reason=row[1],
                                          path_taken=row[2], checks=db.fetch_checks_for_upload(conn, upload_id)))
            continue

        try:
            result = run_fastag_sequence(
                conn, client, image_path=image_path, claimed_vrn=manifest_row["claimed_vrn"],
                claimed_barcode=manifest_row["claimed_barcode"], claimed_tag_id=manifest_row["claimed_tag_id"],
                claimed_bank_code=manifest_row["claimed_bank_code"], upload_id=upload_id,
                vision_model=vision_model,
            )
        except OpenRouterInsufficientCredits as exc:
            print(f"Stopped: {exc}", file=sys.stderr)
            print("Top up your OpenRouter account before retrying -- this isn't a per-image "
                  "failure, every remaining row will fail the same way until you do.", file=sys.stderr)
            break
        except Exception as exc:  # noqa: BLE001 - one bad row shouldn't kill the whole batch
            print(f"  error: {exc}", file=sys.stderr)
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision="ERROR", reason=str(exc),
                                          path_taken=None, checks=[]))
            continue

        output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                      vision_model=vision_model, decision=result.decision,
                                      reason=result.reason, path_taken=result.path_taken,
                                      checks=db.fetch_checks_for_upload(conn, upload_id)))

    write_excel(output_rows, args.output)

    t = client.totals
    print(f"\n-- wrote {len(output_rows)} row(s) to {args.output} --", file=sys.stderr)
    print(f"-- this run: {t.calls} call(s), {t.prompt_tokens + t.completion_tokens} tokens, "
          f"${t.cost_usd:.5f} --", file=sys.stderr)


if __name__ == "__main__":
    main()
