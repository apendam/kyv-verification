#!/usr/bin/env python3
"""Run a list of front-image uploads through the KYV gate sequence, one by
one, and write the results to an Excel file — one row per image, columns
in this order: the human-provided claim (what's being asked against the
image), then the script's own findings in the same sequential order the
gate sequence checks them, then a literal claimed-vs-detected VRN
comparison, then cost/latency/token totals for that row.

This is a standalone copy of the batch pipeline in openrouter_checks/ (see
that package's own scripts/batch_check.py) — kept self-contained under
batch_pipeline/kyv_batch/ so it's easy to share on its own. Any future fix
or improvement to the core checks needs to be applied in both places; see
this repo's batch_pipeline/README.md.

Manifest columns (CSV or .xlsx, header row required):
    image_path      - required. Path to the image (relative paths are
                       resolved relative to the manifest's own directory).
    vrn             - required. Claimed vehicle registration number.
    make            - required. Claimed manufacturer/make.
    vehicle_type    - required. Claimed vehicle type: "bus" or "truck".
    upload_id       - optional. Auto-generated from the VRN if left blank.
    vision_model    - optional. Per-row model override; falls back to
                       --model, then config.DEFAULT_VISION_MODEL.

Usage:
    python scripts/batch_check.py --manifest uploads.csv --output results.xlsx

    # force a different model for every row that doesn't specify its own:
    python scripts/batch_check.py --manifest uploads.xlsx --output results.xlsx \\
        --model google/gemini-2.5-flash
"""
from __future__ import annotations

import argparse
import csv
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv  # noqa: E402

load_dotenv()

from openpyxl import Workbook  # noqa: E402

from kyv_batch import config, db  # noqa: E402
from kyv_batch.client import OpenRouterClient, OpenRouterInsufficientCredits  # noqa: E402
from kyv_batch.gate_sequence import run_gate_sequence  # noqa: E402

REQUIRED_MANIFEST_COLUMNS = ("image_path", "vrn", "make", "vehicle_type")

# Column order in the output Excel file -- input claim, then the gate
# sequence's own findings in the order the checks actually run, then the
# explicit VRN comparison, then the overall decision and cost/latency totals.
OUTPUT_COLUMNS = [
    "image_path", "upload_id", "claimed_vrn", "claimed_make", "claimed_vehicle_type", "vision_model",
    "detected_vehicle_type", "vehicle_type_match", "is_altered_or_ai_generated", "front_image_reasoning",
    "vrn_plate_readable", "vrn_detected_text", "vrn_fuzzy_outcome", "vrn_edit_distance", "vrn_exact_match",
    "maker_readable", "maker_detected_text", "maker_outcome",
    "duplicate_signal_used", "is_duplicate", "duplicate_match_upload_ids",
    "duplicate_match_scores", "duplicate_match_vrns",
    "final_decision", "final_reason",
    "latency_ms_model_checks", "latency_ms_total",
    "prompt_tokens", "completion_tokens", "cost_usd",
]

_MODEL_CHECK_NAMES = {"front_image_check", "vrn_check", "maker_check"}


def read_manifest(path: str) -> list[dict]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        with open(path, newline="") as f:
            rows = [dict(row) for row in csv.DictReader(f)]
    elif ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() for h in next(rows_iter)]
        rows = [dict(zip(header, row)) for row in rows_iter if any(c is not None for c in row)]
    else:
        raise ValueError(f"unsupported manifest format {ext!r} -- use .csv or .xlsx")

    missing = REQUIRED_MANIFEST_COLUMNS - rows[0].keys() if rows else set()
    if missing:
        raise ValueError(f"manifest is missing required column(s): {sorted(missing)}")
    return rows


def _fmt_score(score: float) -> str:
    return str(int(score)) if float(score).is_integer() else f"{score:.4f}"


def build_row(manifest_row: dict, *, image_path: str, upload_id: str, vision_model: str,
              decision: str, reason: str, checks: list[dict]) -> dict:
    # Later rows win on repeats (a --force re-run) -- checks is already in
    # chronological order, see db.fetch_checks_for_upload.
    by_name: dict[str, dict] = {}
    for c in checks:
        by_name[c["check_name"]] = c

    row: dict = {
        "image_path": image_path,
        "upload_id": upload_id,
        "claimed_vrn": manifest_row["vrn"],
        "claimed_make": manifest_row["make"],
        "claimed_vehicle_type": manifest_row["vehicle_type"],
        "vision_model": vision_model,
    }

    front = by_name.get("front_image_check")
    if front:
        d = front["detail"]
        detected_type = d.get("detected_vehicle_type")
        row["detected_vehicle_type"] = detected_type
        claimed_type = (manifest_row["vehicle_type"] or "").strip().lower()
        row["vehicle_type_match"] = detected_type == claimed_type if detected_type else None
        row["is_altered_or_ai_generated"] = d.get("is_altered_or_ai_generated")
        row["front_image_reasoning"] = d.get("reasoning")
    else:
        row.update(dict.fromkeys(
            ["detected_vehicle_type", "vehicle_type_match", "is_altered_or_ai_generated",
             "front_image_reasoning"]))

    vrn_c = by_name.get("vrn_check")
    if vrn_c:
        d = vrn_c["detail"]
        detected_vrn = (d.get("plate_text") or "").strip()
        row["vrn_plate_readable"] = d.get("plate_readable")
        row["vrn_detected_text"] = detected_vrn
        row["vrn_fuzzy_outcome"] = vrn_c["verdict"]
        row["vrn_edit_distance"] = d.get("match_detail", {}).get("distance")
        claimed_vrn_norm = (manifest_row["vrn"] or "").strip().upper()
        row["vrn_exact_match"] = bool(detected_vrn) and detected_vrn.upper() == claimed_vrn_norm
    else:
        row.update(dict.fromkeys(
            ["vrn_plate_readable", "vrn_detected_text", "vrn_fuzzy_outcome",
             "vrn_edit_distance", "vrn_exact_match"]))

    maker_c = by_name.get("maker_check")
    if maker_c:
        d = maker_c["detail"]
        row["maker_readable"] = d.get("maker_readable")
        row["maker_detected_text"] = d.get("maker_text")
        row["maker_outcome"] = maker_c["verdict"]
    else:
        row.update(dict.fromkeys(["maker_readable", "maker_detected_text", "maker_outcome"]))

    dup_c = by_name.get("duplicate_check")
    if dup_c:
        d = dup_c["detail"]
        matches = d.get("matches", [])
        row["duplicate_signal_used"] = d.get("signal")
        row["is_duplicate"] = dup_c["verdict"] == "duplicate"
        row["duplicate_match_upload_ids"] = ", ".join(m["upload_id"] for m in matches)
        row["duplicate_match_scores"] = ", ".join(_fmt_score(m["score"]) for m in matches)
        row["duplicate_match_vrns"] = ", ".join(m.get("claimed_vrn") or "—" for m in matches)
    else:
        row.update(dict.fromkeys(
            ["duplicate_signal_used", "is_duplicate",
             "duplicate_match_upload_ids", "duplicate_match_scores", "duplicate_match_vrns"]))

    row["final_decision"] = decision
    row["final_reason"] = reason

    model_latency = sum(c["latency_ms"] for c in checks if c["check_name"] in _MODEL_CHECK_NAMES)
    duplicate_latency = sum(c["latency_ms"] for c in checks if c["check_name"] == "duplicate_check")
    row["latency_ms_model_checks"] = model_latency
    row["latency_ms_total"] = model_latency + duplicate_latency
    row["prompt_tokens"] = sum(c["prompt_tokens"] for c in checks)
    row["completion_tokens"] = sum(c["completion_tokens"] for c in checks)
    row["cost_usd"] = sum(c["cost_usd"] for c in checks)

    return row


def write_excel(rows: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Results"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(col) for col in OUTPUT_COLUMNS])
    wb.save(output_path)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--manifest", required=True, help="CSV or .xlsx file listing images to check.")
    ap.add_argument("--output", required=True, help="Excel file to write results to.")
    ap.add_argument("--model", default=config.DEFAULT_VISION_MODEL,
                     help=f"Default vision model for rows without their own 'vision_model' "
                          f"column value (default: {config.DEFAULT_VISION_MODEL}).")
    ap.add_argument("--db", default=str(config.DEFAULT_DB_PATH), help="SQLite file path.")
    ap.add_argument("--force", action="store_true",
                     help="Re-run rows whose upload_id was already checked (by default, skipped).")
    args = ap.parse_args()

    manifest_rows = read_manifest(args.manifest)
    manifest_dir = Path(args.manifest).resolve().parent
    conn = db.connect(args.db)
    client = OpenRouterClient()

    output_rows: list[dict] = []
    for i, manifest_row in enumerate(manifest_rows, start=1):
        image_path = manifest_row["image_path"].strip()
        if not Path(image_path).is_absolute():
            candidate = manifest_dir / image_path
            image_path = str(candidate if candidate.is_file() else Path(image_path))

        vrn = (manifest_row["vrn"] or "").strip()
        vehicle_type = (manifest_row["vehicle_type"] or "").strip()
        upload_id = (manifest_row.get("upload_id") or "").strip() or f"batch-{vrn}-{time.time_ns()}"
        vision_model = (manifest_row.get("vision_model") or "").strip() or args.model

        print(f"[{i}/{len(manifest_rows)}] {upload_id} <- {image_path}", file=sys.stderr)

        if not Path(image_path).is_file():
            print(f"  skipped: image not found: {image_path}", file=sys.stderr)
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision="SKIPPED",
                                          reason="image not found", checks=[]))
            continue

        if not args.force and db.already_checked(conn, upload_id):
            print(f"  skipped: '{upload_id}' already checked (pass --force to re-run)", file=sys.stderr)
            row = conn.execute("SELECT decision, reason FROM results WHERE upload_id = ?",
                                (upload_id,)).fetchone()
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision=row[0], reason=row[1],
                                          checks=db.fetch_checks_for_upload(conn, upload_id)))
            continue

        try:
            result = run_gate_sequence(
                conn, client, image_path=image_path, claimed_vrn=vrn, claimed_make=manifest_row["make"],
                claimed_vehicle_type=vehicle_type, upload_id=upload_id, vision_model=vision_model,
            )
        except OpenRouterInsufficientCredits as exc:
            print(f"Stopped: {exc}", file=sys.stderr)
            print("Top up your OpenRouter account before retrying -- this isn't a per-image "
                  "failure, every remaining row will fail the same way until you do.", file=sys.stderr)
            break
        except Exception as exc:  # noqa: BLE001 - one bad row shouldn't kill the whole batch
            print(f"  error: {exc}", file=sys.stderr)
            output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                          vision_model=vision_model, decision="ERROR", reason=str(exc),
                                          checks=[]))
            continue

        output_rows.append(build_row(manifest_row, image_path=image_path, upload_id=upload_id,
                                      vision_model=vision_model, decision=result.decision,
                                      reason=result.reason,
                                      checks=db.fetch_checks_for_upload(conn, upload_id)))

    write_excel(output_rows, args.output)

    t = client.totals
    print(f"\n-- wrote {len(output_rows)} row(s) to {args.output} --", file=sys.stderr)
    print(f"-- this run: {t.calls} call(s), {t.prompt_tokens + t.completion_tokens} tokens, "
          f"${t.cost_usd:.5f} --", file=sys.stderr)


if __name__ == "__main__":
    main()
